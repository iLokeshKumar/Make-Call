import sys
import openpyxl
import re

wb = openpyxl.load_workbook('Order Process 2026-27.xlsx')
ws = wb['TNKL PO Sheet']

MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
MONTH_NUMS = {str(i).zfill(2): m for i, m in enumerate(MONTHS, 1)}

def parse_month(date_val):
    if date_val is None:
        return None
    if isinstance(date_val, str) and date_val.startswith('='):
        return None
    if isinstance(date_val, str):
        for sep in ['-', '/']:
            parts = date_val.split(sep)
            if len(parts) == 3:
                month_num = parts[1].zfill(2)
                if month_num.isdigit() and 1 <= int(month_num) <= 12:
                    return MONTH_NUMS[month_num]
    if hasattr(date_val, 'month'):
        return MONTHS[date_val.month - 1]
    return None

def parse_po_value(val):
    if val is None:
        return None
    if isinstance(val, str) and val.startswith('='):
        return None
    try:
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val)
        s = s.replace(',', '')
        s = re.sub(r'[^0-9.\-]', '', s)
        if s:
            return float(s)
    except:
        pass
    return None

print("=" * 120)
print("COMPLETE AUDIT: TNKL PO Sheet - ALL ROWS")
print("=" * 120)
print(f"{'Row':>4} {'SNO':>4} {'SPOC':20s} {'Date(ColF)':20s} {'PO Date(ColL)':20s} {'Product':50s} {'PO Value':20s} {'Month':10s} {'Included?':10s}")
print("-" * 120)

included_count = 0
excluded_count = 0
aswini_included = 0
aswini_excluded = 0
aswini_details = []

for row in range(2, ws.max_row + 1):
    sno = ws.cell(row=row, column=1).value
    if sno is None:
        continue
    date_val = ws.cell(row=row, column=6).value
    po_date_val = ws.cell(row=row, column=12).value
    spoc = ws.cell(row=row, column=7).value
    product = ws.cell(row=row, column=10).value
    po_value = ws.cell(row=row, column=13).value

    month = parse_month(date_val)
    po_parsed = parse_po_value(po_value)
    
    included = month is not None and po_parsed is not None
    
    prod_str = str(product).replace('\\n', ' ').strip() if product else ''
    if len(prod_str) > 45:
        prod_str = prod_str[:42] + '...'
    
    po_str = str(po_value) if po_value is not None else ''
    if len(po_str) > 18:
        po_str = po_str[:15] + '...'
    
    month_str = month if month else 'NO DATE'
    incl_str = 'YES' if included else 'NO'
    
    print(f"{row:>4} {str(sno):>4} {str(spoc or ''):20s} {str(date_val):20s} {str(po_date_val):20s} {prod_str:50s} {po_str:20s} {month_str:10s} {incl_str:10s}")
    
    if spoc and 'aswini' in str(spoc).lower():
        if included:
            aswini_included += 1
        else:
            aswini_excluded += 1
            reason = []
            if month is None:
                reason.append(f"NO_DATE(date={date_val})")
            if po_parsed is None:
                reason.append(f"BAD_PO_VALUE({po_value})")
            aswini_details.append((row, sno, date_val, po_value, month, po_parsed, ' | '.join(reason)))

print("-" * 120)
print(f"\nSUMMARY: Included={included_count}, Excluded={excluded_count}")

print(f"\n{'=' * 120}")
print("ASWINI.V AUDIT")
print(f"{'=' * 120}")
print(f"Included: {aswini_included}")
print(f"Excluded: {aswini_excluded}")

print(f"\n--- Excluded aswini.v orders (missing or problematic): ---")
for row, sno, dt, pv, month, parsed, reason in aswini_details:
    print(f"  Row{row} SNO={sno}: {reason}")
    print(f"    Date={dt}, PO Value={pv}, parsed={parsed}")

print(f"\n{'=' * 120}")
print("CSD PO SHEET - ALL ROWS AUDIT")
print(f"{'=' * 120}")

ws2 = wb['CSD PO Sheet']
csd_total = 0
csd_included = 0
csd_excluded = 0

for row in range(2, ws2.max_row + 1):
    sno = ws2.cell(row=row, column=1).value
    if sno is None:
        continue
    date_val = ws2.cell(row=row, column=5).value
    isr_gsm = ws2.cell(row=row, column=6).value
    product = ws2.cell(row=row, column=12).value
    po_value = ws2.cell(row=row, column=16).value
    
    month = parse_month(date_val)
    po_parsed = parse_po_value(po_value)
    included = month is not None and po_parsed is not None
    
    csd_total += 1
    if included:
        csd_included += 1
    else:
        csd_excluded += 1
        reason = []
        if month is None:
            reason.append(f"NO_DATE({date_val})")
        if po_parsed is None:
            reason.append(f"BAD_PO({po_value})")
        print(f"  Row{row} SNO={sno} | {isr_gsm} | Date={date_val} | PO={po_value} | {' | '.join(reason)}")

print(f"\nCSD Summary: Total={csd_total}, Included={csd_included}, Excluded={csd_excluded}")
