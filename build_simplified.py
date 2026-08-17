import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

wb = openpyxl.load_workbook('Order Process 2026-27.xlsx')

MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
MONTH_NUMS = {str(i).zfill(2): m for i, m in enumerate(MONTHS, 1)}

import re

def parse_month(date_val):
    """Extract month name from various date formats"""
    if date_val is None:
        return None
    if isinstance(date_val, str) and date_val.startswith('='):
        return None
    if isinstance(date_val, str):
        for sep in ['-', '/']:
            parts = date_val.split(sep)
            if len(parts) == 3:
                month_num = parts[1].zfill(2)
                if month_num.isdigit() and 1 <= int(month_num) <= 12:
                    return MONTH_NUMS[month_num]
    if hasattr(date_val, 'month'):
        return MONTHS[date_val.month - 1]
    return None

def safe_parse_po_value(val):
    """Safely parse PO value handling: Indian comma format (1,53,164), currency symbols, etc."""
    if val is None:
        return None
    if isinstance(val, str) and val.startswith('='):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        # Strip commas, currency symbols, and any non-numeric chars except dot and minus
        s = str(val).replace(',', '')
        s = re.sub(r'[^0-9.\-]', '', s)
        if s:
            return float(s)
    except (ValueError, TypeError):
        pass
    return None

def build_pivot(data):
    pivot = defaultdict(lambda: defaultdict(float))
    all_months = set()
    for keys, month, val in data:
        if month is None or month == '':
            continue
        po_parsed = safe_parse_po_value(val)
        if po_parsed is None:
            continue
        pivot[keys][month] += po_parsed
        all_months.add(month)
    return pivot, sorted(all_months, key=lambda m: list(MONTHS).index(m))

# TNKL PO SHEET
ws_tnkl = wb['TNKL PO Sheet']
tnkl_data = []
for row in range(2, ws_tnkl.max_row + 1):
    sno = ws_tnkl.cell(row=row, column=1).value
    if sno is None:
        continue
    date_val = ws_tnkl.cell(row=row, column=6).value      # Col F = Date (order date)
    po_date_val = ws_tnkl.cell(row=row, column=12).value  # Col L = PO Date (fallback)
    spoc = ws_tnkl.cell(row=row, column=7).value
    product = ws_tnkl.cell(row=row, column=10).value
    po_value = ws_tnkl.cell(row=row, column=13).value
    if isinstance(po_value, str) and po_value.startswith('='):
        continue
    # Try main date col first, fall back to PO Date col
    month = parse_month(date_val)
    if month is None:
        month = parse_month(po_date_val)
    if month:
        prod_str = str(product).replace('\n', ' ').strip() if product else ''
        if len(prod_str) > 80:
            prod_str = prod_str[:77] + '...'
        tnkl_data.append(((spoc or 'N/A', prod_str), month, po_value))

tnkl_pivot, tnkl_months = build_pivot(tnkl_data)

# CSD PO SHEET
ws_csd = wb['CSD PO Sheet']
csd_data = []
for row in range(2, ws_csd.max_row + 1):
    sno = ws_csd.cell(row=row, column=1).value
    if sno is None:
        continue
    date_val = ws_csd.cell(row=row, column=5).value   # Col E = Date
    isr_gsm = ws_csd.cell(row=row, column=6).value     # Col F = ISR / GSM
    product = ws_csd.cell(row=row, column=12).value    # Col L = PRODUCT
    po_value = ws_csd.cell(row=row, column=16).value   # Col P = PO Value
    if isinstance(po_value, str) and po_value.startswith('='):
        continue
    month = parse_month(date_val)
    if month:
        prod_str = str(product).replace('\n', ' ').strip() if product else ''
        if len(prod_str) > 80:
            prod_str = prod_str[:77] + '...'
        csd_data.append(((isr_gsm or 'N/A', prod_str), month, po_value))

csd_pivot, csd_months = build_pivot(csd_data)

# CREATE NEW WORKBOOK
wb2 = openpyxl.Workbook()

header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
currency_format = '#,##0.00'

def write_sheet(ws, title, pivot_data, sorted_months, key_cols):
    n_keys = len(key_cols)
    n_months = len(sorted_months)
    total_cols = n_keys + n_months + 1

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    tc = ws.cell(row=1, column=1, value=title)
    tc.font = Font(bold=True, size=14, color='2F5496')
    tc.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30

    # Headers
    headers = key_cols + [f"{m}'26" for m in sorted_months] + ['Total']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align
        c.border = thin_border
    ws.row_dimensions[2].height = 25

    # Data
    row_num = 3
    sorted_keys = sorted(pivot_data.keys(), key=lambda k: (k[0].lower(), k[1].lower()))
    grand_total = 0.0
    month_totals = defaultdict(float)

    for keys in sorted_keys:
        months_dict = pivot_data[keys]
        key_vals = list(keys)
        while len(key_vals) < n_keys:
            key_vals.append('')

        for ci, val in enumerate(key_vals, 1):
            c = ws.cell(row=row_num, column=ci, value=val)
            c.border = thin_border
            c.alignment = Alignment(wrap_text=True, vertical='top')

        row_total = 0.0
        for mi, m in enumerate(sorted_months):
            val = months_dict.get(m, 0.0)
            if val > 0:
                c = ws.cell(row=row_num, column=n_keys + mi + 1, value=round(val, 2))
                c.number_format = currency_format
                c.border = thin_border
                c.alignment = Alignment(horizontal='right')
                row_total += val
                month_totals[m] += val

        tc = ws.cell(row=row_num, column=total_cols, value=round(row_total, 2))
        tc.number_format = currency_format
        tc.font = Font(bold=True)
        tc.border = thin_border
        tc.alignment = Alignment(horizontal='right')
        grand_total += row_total
        ws.row_dimensions[row_num].height = 22
        row_num += 1

    # Summary row
    row_num += 1
    sum_label = ws.cell(row=row_num, column=1, value='MONTH TOTALS')
    sum_label.font = Font(bold=True, color='FFFFFF', size=12)
    sum_label.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    sum_label.alignment = center_align
    sum_label.border = thin_border
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=n_keys)

    blue_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    for mi, m in enumerate(sorted_months):
        c = ws.cell(row=row_num, column=n_keys + mi + 1, value=round(month_totals[m], 2))
        c.number_format = currency_format
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = blue_fill
        c.alignment = Alignment(horizontal='right')
        c.border = thin_border

    c = ws.cell(row=row_num, column=total_cols, value=round(grand_total, 2))
    c.number_format = currency_format
    c.font = Font(bold=True, color='FFFFFF', size=12)
    c.fill = blue_fill
    c.alignment = Alignment(horizontal='right')
    c.border = thin_border

    # Count row
    row_num += 1
    cnt = ws.cell(row=row_num, column=1, value=f'Total Orders: {len(sorted_keys)}')
    cnt.font = Font(italic=True, color='666666')
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=total_cols)

    # Column widths
    for i in range(1, n_keys + 1):
        ws.column_dimensions[get_column_letter(i)].width = 28
    for i in range(n_keys + 1, total_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14

    ws.sheet_properties.tabColor = '2F5496'

# Sheet 1: TNKL PO
ws1 = wb2.active
ws1.title = 'TNKL PO Sheet'
write_sheet(ws1, 'TNKL PO Summary — 2026-27', tnkl_pivot, tnkl_months, ['SPOC Name', 'Product'])

# Sheet 2: CSD PO
ws2 = wb2.create_sheet('CSD PO Sheet')
write_sheet(ws2, 'CSD PO Summary — 2026-27', csd_pivot, csd_months, ['ISR / GSM', 'Product'])

wb2.save('Order_Process_Simplified.xlsx')
print('SAVED: Order_Process_Simplified.xlsx')
